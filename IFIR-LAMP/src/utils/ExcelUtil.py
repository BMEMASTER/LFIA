# -*- coding: utf-8 -*-
""" Excel 导入/导出API """
from openpyxl import Workbook, load_workbook, styles
import numpy as np


def exportData(filename, data: list):
    """ 导出数据到excel """
    dataArr = np.array(data)
    rows, cols = dataArr.shape
    wb = Workbook()
    sheet = wb.active
    for row in range(rows):
        for col in range(cols):
            sheet.cell(row + 1, col + 1).alignment = styles.Alignment(horizontal='center', vertical='center')
            sheet.cell(row + 1, col + 1).value = dataArr[row, col]
    wb.save(filename)


def importData(fileName):
    """ 读取excel中的数据 """
    recordList = []
    # 加载excel文件
    wb = load_workbook(fileName)
    st = wb.active
    # 获取行和列
    rows = st.max_row
    cols = st.max_column
    for row in range(1, rows + 1):
        record = []
        for col in range(1, cols + 1):
            record.append(st.cell(row, col).value)
        recordList.append(tuple(record))
    return recordList
