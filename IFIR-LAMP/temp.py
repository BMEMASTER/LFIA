# -*- coding: utf-8 -*-
import cv2 as cv
import numpy as np
from openpyxl import Workbook
import matplotlib.pyplot as plt


def temp():
    """
        临时存放代码
    """
    # cv.imshow("", gray)
    # chs = cv.split(srcImg)
    # rCh = chs[2]    # 红色通道
    # cv.imshow("rCh", rCh)
    # # 归一化
    # norimg = np.zeros(rCh.shape, dtype=np.uint8)
    # cv.normalize(rCh, norimg, 0, 255, cv.NORM_MINMAX)
    # cv.imshow("norimg", norimg)
    # # 保存归一化的图片
    # cv.imwrite(path + "normalize_" + imgName, norimg)


def createProjectionImg(values, shape):
    """
        绘制投影图
    """
    if len(shape) != 2 or len(values) != shape[1]:
        return None
    rows = 256
    cols = shape[1]
    dst = np.zeros((rows, shape[1]), dtype=np.uint8)
    for i in range(cols):
        row = 255 - np.uint8(values[i])
        dst[row, i] = row
    return dst


def save2Excel(data, name, fileName):
    """
        将数据保存到excel中
    """
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "Kind"
    sheet["A2"] = name
    for i in range(len(data)):
        sheet.cell(1, i + 2).value = i
        sheet.cell(2, i + 2).value = data[i]
    wb.save(fileName)


if __name__ == "__main__":
    path = r"C:\Users\xhc\Desktop"       # 文件路径
    imgName = "22222"    # 图片名称
    srcImg = cv.imread(r"{}\dst_1.png".format(path))
    gray = cv.cvtColor(srcImg, cv.COLOR_BGR2GRAY)
    # 求每列的平均值
    colPixAve = np.mean(gray, 0)
    # 绘制投影图
    plt.plot(colPixAve)
    plt.show()
    # save2Excel(colPixAve, imgName, r"{}\{}.xlsx".format(path, imgName))
    # cv.waitKey()
    dst = np.zeros(gray.shape, dtype=gray.dtype)
    cv.threshold(gray, 0, 255, cv.THRESH_BINARY | cv.THRESH_TRIANGLE, dst)
    cv.imwrite("{}/222.png".format(path), dst)

