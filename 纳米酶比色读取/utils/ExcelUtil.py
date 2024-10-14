# -*- coding: utf-8 -*-
""" Excel 导入/导出API """
from openpyxl import Workbook, load_workbook, styles
import numpy as np


def exportData(filename, data: list):
    """ 导出数据到excel """
    dataArr = np.array(data)
    rows, cols = dataArr.shape
    wb = Workbook()
    ws = wb.active
    # 合并单元格
    ws.merge_cells(start_column=1, end_column=3, start_row=1, end_row=1)
    # 插入时间
    ws.cell(1, 1).alignment = styles.Alignment(horizontal='center', vertical='center')
    ws.cell(1, 1).value = dataArr[0, 0]
    for row in range(1, rows):
        for col in range(cols):
            ws.cell(row + 1, col + 1).alignment = styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row + 1, col + 1).value = dataArr[row, col]
    wb.save(filename)


def appendData(filename, data: list):
    # 创建工作薄
    wb = load_workbook(filename)
    ws = wb.active
    # 获取表格中已经有的行数，数据从下一行开始插入
    beginRow = ws.max_row
    beginRow = beginRow + 2 if beginRow > 1 else 1
    # 保存数据
    dataArr = np.array(data)
    rows, cols = dataArr.shape
    # 合并单元格
    ws.merge_cells(start_column=1, end_column=3, start_row=beginRow, end_row=beginRow)
    # 插入时间
    ws.cell(beginRow, 1).alignment = styles.Alignment(horizontal='center', vertical='center')
    ws.cell(beginRow, 1).value = dataArr[0, 0]
    # 插入数据
    for row in range(1, rows):
        for col in range(cols):
            ws.cell(row + beginRow, col + 1).alignment = styles.Alignment(horizontal='center', vertical='center')
            ws.cell(row + beginRow, col + 1).value = dataArr[row, col]
    wb.save(filename)


def importData(fileName):
    """ 读取excel中的数据 """
    recordList = []
    # 加载excel文件
    wb = load_workbook(fileName)
    st = wb.active
    # 获取行和列
    rows = st.max_row
    cols = st.max_column
    for row in range(1, rows + 1):
        record = []
        for col in range(1, cols + 1):
            record.append(st.cell(row, col).value)
        recordList.append(tuple(record))
    return recordList


if __name__ == "__main__":
    wb = load_workbook('C:/Users/xhc/Desktop/1.xlsx')
    ws = wb.active
    print(ws.max_row)
    print(ws.max_column)

